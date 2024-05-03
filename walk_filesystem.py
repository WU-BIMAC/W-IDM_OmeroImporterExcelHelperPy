import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# maximum file name length
MAX_NAME = 255

def truncate_name(dirs, max=MAX_NAME):
    """Construct new file name from the end to the front until it exceeds MAX_NAME.

     Args:
        dirs: list of directory names ["tag1", "tag2", ..., "tagN", file].
        max: integer maximum length for file name.
     Returns:
        String: file name shorter than MAX_NAME
    """
    new_filename = "\\".join(dirs)
    length = len(new_filename)
    num_split = 1
    while length > max:
        # split off front num_split directories
        split = new_filename.split("\\", num_split)
        length = len(split[-1])
        num_split += 1
    new_filename = split[-1]
    return new_filename

# WARNING: : str.count() is case sensitive
def image_walk(root, extensions):
    """Navigates each file and folder beginning at the root folder and checks for image files.

    Args:
        root: path of the root project folder.
        extensions: list of string image file extensions.
    Yields:
        Yields file name and file path of each image file found.
    Raises:
    """
    for dirpath, dirnames, files in os.walk(root, topdown=True):
        for file in files:
            filename, extension = os.path.splitext(file)
            if extension in extensions:
                filepath = os.path.join(dirpath, file)
                # get accompanying json file
                json = None
                for f in os.listdir(dirpath):
                    name, ext = os.path.splitext(f)
                    if (ext == ".json") and (name.lower().count(file.lower()) > 0):
                        json = f
                yield (file, filepath, json)

def create_DataFrame(root, excel, columns=["File Name","New File Name","File Path","MMA File Path","Tags"]):
    """Create a DataFrame containing image file info for the root directory.

    DataFrame will contain file name, new file name constructed by replacing backslashes in filepath
    with underscores, file path, MMA file path for the accompanying json file path, and tags which are the root directories subdirectories seperated by hashtags.

     Args:
        root: path to a dataset directory continaing subdirectories and image files.
        excel: string excel file name to write to.
        columns: column names used to create the DataFrame and CSV output.
     Returns:
        A DataFrame containing information written to the output csv file.
     Catches:
        ValueError: catches value error from image_walk() if encountering a path longer than MAX_NAME characters
    """
    df = pd.DataFrame(columns=columns)
    extensions = get_extensions(excel)
    walk = image_walk(root, extensions)

    for file, filepath, json in walk:
        # seperate the directories in the path from the ending file
        dirpath = os.path.dirname(filepath)

        # remove the root part of the dirname
        dirpath = dirpath.replace(root, "", 1)

        # split dirpath into list of directories
        dirs = dirpath.split("\\")

        # remove empty string left by the leading "\" after splitting
        dirs.remove("")

        # seperate tags with #
        tags = "#".join(dirs)

        dirs.append(file)

        # join directories with "_"
        new_filename = "_".join(dirs)

        # truncate file name if too long
        if len(new_filename) > MAX_NAME:
            new_filename = truncate_name(dirs)

        # add row to the DataFrame
        row = pd.DataFrame([[file, new_filename, filepath, json, tags]], columns=columns)
        df = df.append(row, ignore_index=True)

    df.sort_values(by="File Name", inplace=True)
    display(df)
    return df

# WARNING: : if xlsx keep_vba must be False, if xlsm keep_vba must be true, otherwise file will be corrupted
def write_excel(file, df, sheet_number=2, start_row=14, start_col=1):
    """Write DataFrame to an existing excel file on a specific sheet starting at a specific cell.

     Args:
        file: string file name.
        df: DataFrame table to write.
        sheet_number: integer zero-indexed index number of the sheet to write on.
        start_row: integer one-indexed index number of row part of cell.
        start_col: integer one-indexed index number of column part of cell.
     Returns:
        None
     Raises:
    """
    name, ext = os.path.splitext(file)
    if (ext == ".xlsx") or (ext == ".xltx"): # NOTE: only two excel files which cannot have macros
        wb = load_workbook(file, read_only=False, keep_vba=False)
    else:
        wb = load_workbook(file, read_only=False, keep_vba=True)
    sheetname = wb.sheetnames[sheet_number]
    sheet = wb[sheetname]
    rows = dataframe_to_rows(df, header=False, index=False)
    for r_idx, row in enumerate(rows, start_row):
        for c_idx, value in enumerate(row, start_col):
             sheet.cell(row=r_idx, column=c_idx).value = value
    wb.save(file)

# NOTE: I get the sheet name from the sheet index before selecting the sheet because of a naming bug where the unicode character U+0399 capital Greek Iota is present instead of U+0049 capital Roman I
def get_extensions(file, sheet_number=2, cell_number="B10"):
    """Read an excel file and extract image file extenions from a specific cell.

     Args:
        file: excel file.
        sheet_number: zero-indexed index of Image-List excel worksheet.
        cell_number: string row and column of excele worksheet cell containing list of image file extensions.
     Returns:
        A list of image file extensions.
     Raises:
        UserWarning: Data Validation extension is not supported and will be removed.
    """
    wb = load_workbook(file, read_only=True)
    sheetname = wb.sheetnames[sheet_number]
    sheet = wb[sheetname]
    cell = sheet[cell_number]
    extensions = cell.value.split(",")
    wb.close()
    return extensions

# TODO: pass excel file name this is being called from as argument to main
# TODO: will need to find dataset folder name from cell in excel file
def main():
    """Main method for running the script.

    Loops through current directory to find each subdirectory which should be for a dataset of images,
    and calls create_DataFrame() to output a csv for each dataset.

     Args:
     Returns:
        None
     Raises:
    """
    # root = "." # NOTE: relative path, can possibly replace full root with "."
    cwd = os.getcwd()
    root = os.path.splitdrive(cwd)[1]

    # # name csv file after dataset folder
    # csv_name = os.path.basename(root)
    # df.to_csv("{}.csv".format(csv_name), index=False)

    df = create_DataFrame(os.path.join(root, "Dataset 1"), "2022 Pazour Import Test Project 3#Dataset 1.xlsx")
    write_excel("2022 Pazour Import Test Project 3#Dataset 1.xlsx", df)

    # df = create_DataFrame(os.path.join(root, "Dataset 2"), "2022 Pazour Import Test Project 3#Dataset 2.xlsx")
    #
    # df = create_DataFrame(os.path.join(root, "Dataset_Name01"), "Project_Name0_Dataset_Name01.xlsm")
    #
    # df = create_DataFrame(os.path.join(root, "Dataset_Name11"), "Project_Name1_Dataset_Name11.xlsm")
    #
    # df = create_DataFrame(os.path.join(root, "Dataset_Name21"), "Project_Name2_Dataset_Name21.xlsm")
    #
    # df = create_DataFrame(os.path.join(root, "Dataset_Name22"), "Project_Name2_Dataset_Name22.xlsm")

if __name__ == "__main__":
    main()
